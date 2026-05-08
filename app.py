import os
import gc
import zipfile
from xhtml2pdf import pisa
from flask import Flask, render_template, request, send_file, jsonify
from werkzeug.exceptions import HTTPException
import pandas as pd
import numpy as np
from datetime import datetime
import traceback
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'super_secret_key'
# Removed MAX_CONTENT_LENGTH limit to allow files of any size
# app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024

def safe_read_excel(path, **kwargs):
    """Read an Excel file using openpyxl engine."""
    kwargs['engine'] = 'openpyxl'
    return pd.read_excel(path, **kwargs)

# On Render (and similar platforms) only /tmp is guaranteed writable.
# Locally we use the project-relative folders as before.
if os.environ.get('RENDER'):
    UPLOAD_FOLDER = '/tmp/uploads'
    OUTPUT_FOLDER = '/tmp/output'
else:
    UPLOAD_FOLDER = 'uploads'
    OUTPUT_FOLDER = 'output'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
TOTAL_FILL   = PatternFill("solid", fgColor="D6E4F0")
ALT_ROW_FILL = PatternFill("solid", fgColor="EBF2FA")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TOTAL_FONT  = Font(name="Calibri", bold=True, color="1F3864", size=11)
BODY_FONT   = Font(name="Calibri", size=10)
TITLE_FONT  = Font(name="Calibri", bold=True, size=13, color="1F3864")

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
TOTAL_BORDER = Border(
    left=Side(style="medium", color="1F3864"),
    right=Side(style="medium", color="1F3864"),
    top=Side(style="medium", color="1F3864"),
    bottom=Side(style="medium", color="1F3864"),
)

PCT_FORMAT    = '0%'
RATING_FORMAT = '0.0'
INT_FORMAT    = '#,##0'


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
def _style_sheet(ws, title_label: str,
                 text_cols: set, pct_cols: set, rating_cols: set, int_cols: set):
    """Insert title + date rows then apply professional formatting."""
    ws.insert_rows(1)
    ws.insert_rows(1)

    title_cell = ws.cell(row=1, column=1, value=title_label)
    title_cell.font      = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    date_cell = ws.cell(row=2, column=1,
                        value=f"Generated: {datetime.now().strftime('%d %b %Y  %I:%M %p')}")
    date_cell.font      = Font(name="Calibri", italic=True, size=9, color="7F7F7F")
    date_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14

    header_row = 3
    max_col = ws.max_column
    max_row = ws.max_row

    # Auto-fit column widths
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                col_widths[cell.column] = max(
                    col_widths.get(cell.column, 8),
                    min(len(str(cell.value)) + 4, 40)
                )
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Header row
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 30

    # Data rows (everything except the last = Total row)
    last_data_row = max_row
    for row_idx in range(header_row + 1, last_data_row):
        fill = ALT_ROW_FILL if (row_idx % 2 == 0) else WHITE_FILL
        for col_idx in range(1, max_col + 1):
            cell        = ws.cell(row=row_idx, column=col_idx)
            cell.fill   = fill
            cell.font   = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left" if col_idx in text_cols else "right",
                vertical="center"
            )
            if col_idx in pct_cols:
                cell.number_format = PCT_FORMAT
            elif col_idx in rating_cols:
                cell.number_format = RATING_FORMAT
            elif col_idx in int_cols:
                cell.number_format = INT_FORMAT
        ws.row_dimensions[row_idx].height = 18

    # Total row
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=last_data_row, column=col_idx)
        cell.fill      = TOTAL_FILL
        cell.font      = TOTAL_FONT
        cell.border    = TOTAL_BORDER
        cell.alignment = Alignment(
            horizontal="left" if col_idx in text_cols else "right",
            vertical="center"
        )
        if col_idx in pct_cols:
            cell.number_format = PCT_FORMAT
        elif col_idx in rating_cols:
            cell.number_format = RATING_FORMAT
        elif col_idx in int_cols:
            cell.number_format = INT_FORMAT
    ws.row_dimensions[last_data_row].height = 22

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _write_styled_sheet(writer, df: pd.DataFrame, sheet_name: str,
                        title_label: str, text_col_names: list,
                        pct_col_names: list, rating_col_names: list,
                        int_col_names: list):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    cols = list(df.columns)
    def _idx(names):
        return {cols.index(n) + 1 for n in names if n in cols}

    _style_sheet(ws, title_label,
                 _idx(text_col_names), _idx(pct_col_names),
                 _idx(rating_col_names), _idx(int_col_names))


def _add_total_row(df: pd.DataFrame, label_col: str,
                   sum_cols: list, avg_cols: list,
                   pct_cols_raw: list) -> pd.DataFrame:
    """Append a TOTAL row; % is recalculated from sums, not averaged."""
    total = {c: '' for c in df.columns}
    total[label_col] = 'TOTAL'

    for c in sum_cols:
        if c in df.columns:
            total[c] = df[c].sum()

    for c in avg_cols:
        if c in df.columns:
            valid = df[c][df[c] > 0]
            total[c] = round(valid.mean(), 1) if len(valid) > 0 else 0.0

    for pct_col, fb_col, bill_col in pct_cols_raw:
        tb = total.get(bill_col, 0)
        tf = total.get(fb_col, 0)
        total[pct_col] = (tf / tb) if tb > 0 else 0.0

    return pd.concat([df, pd.DataFrame([total])], ignore_index=True)


# ---------------------------------------------------------------------------
# Load Feedback Data Helper
# ---------------------------------------------------------------------------
def load_feedback_auto(fb_path):
    """Robustly load feedback Excel, handling potential header offsets."""
    try:
        df = safe_read_excel(fb_path)
        check_cols = [str(c).strip().lower() for c in df.columns]
        if 'rating' in check_cols or 'slno' in check_cols or 'branch name' in check_cols:
            df.columns = df.columns.str.strip()
            return df

        df_preview = safe_read_excel(fb_path, header=None, nrows=20)
        header_idx = 0
        for i in range(len(df_preview)):
            row_vals = df_preview.iloc[i].astype(str).str.strip().str.lower().tolist()
            if 'rating' in row_vals or 'branch name' in row_vals or 'slno' in row_vals:
                header_idx = i
                break

        df_actual = safe_read_excel(fb_path, header=header_idx)
        df_actual.columns = df_actual.columns.str.strip()
        return df_actual
    except Exception:
        df = safe_read_excel(fb_path)
        df.columns = df.columns.str.strip()
        return df


def load_sales_auto(sales_path):
    """Robustly load sales Excel, handling potential header offsets and sheet names."""
    try:
        df = safe_read_excel(sales_path, sheet_name='Detailed Sales Report')
        check_cols = [str(c).strip().lower() for c in df.columns]
        if 'staff code' in check_cols or 'customer mobile' in check_cols or 'branch' in check_cols:
            df.columns = df.columns.str.strip()
            return df
    except Exception:
        pass
        
    try:
        df_preview = safe_read_excel(sales_path, sheet_name='Detailed Sales Report', header=None, nrows=20)
        sheet_kwargs = {'sheet_name': 'Detailed Sales Report'}
    except Exception:
        df_preview = safe_read_excel(sales_path, header=None, nrows=20)
        sheet_kwargs = {}
        
    header_idx = 0
    for i in range(len(df_preview)):
        row_vals = df_preview.iloc[i].astype(str).str.strip().str.lower().tolist()
        if 'staff code' in row_vals or 'customer mobile' in row_vals or 'branch' in row_vals:
            header_idx = i
            break
            
    try:
        df_actual = safe_read_excel(sales_path, header=header_idx, **sheet_kwargs)
    except Exception:
        df_actual = safe_read_excel(sales_path, header=0)
        
    df_actual.columns = df_actual.columns.str.strip()
    return df_actual


# ---------------------------------------------------------------------------
# Main processing function
# ---------------------------------------------------------------------------
def process_reports(sales_path, fb_path, output_path):

    # ── Load & clean sales — only the columns we actually need ──────────────
    # Read all columns first to find the right header, then slim down
    df_sales_raw = load_sales_auto(sales_path)

    # Remove 'general' rows from key columns
    for col in ['RBM', 'BDM', 'Staff', 'Branch', 'Category', 'Designation']:
        if col in df_sales_raw.columns:
            df_sales_raw = df_sales_raw[~(df_sales_raw[col].astype(str).str.strip().str.lower() == 'general')]

    # Keep only the columns needed for all downstream operations
    NEEDED_COLS = ['Staff Code', 'Staff', 'Branch', 'RBM', 'BDM', 'Customer Mobile']
    available   = [c for c in NEEDED_COLS if c in df_sales_raw.columns]
    df_sales    = df_sales_raw[available].copy()
    del df_sales_raw
    gc.collect()

    BILL_COL = 'Customer Mobile'
    df_bills = (df_sales.drop_duplicates(subset=[BILL_COL])
                if BILL_COL in df_sales.columns else df_sales.copy())

    staff_bill_cut  = df_bills.groupby('Staff Code').size().reset_index(name='Staff Bills')
    branch_bill_cut = df_bills.groupby('Branch').size().reset_index(name='Branch Bills')
    rbm_bill_cut    = df_bills.groupby('RBM').size().reset_index(name='RBM Bills')
    bdm_bill_cut    = df_bills.groupby('BDM').size().reset_index(name='BDM Bills')
    del df_bills
    gc.collect()

    master_staff = df_sales.drop_duplicates(subset=['Staff Code']).copy()
    del df_sales
    gc.collect()

    staff_to_rbm  = dict(zip(master_staff['Staff'], master_staff['RBM']))
    staff_to_bdm  = dict(zip(master_staff['Staff'], master_staff['BDM']))
    branch_to_rbm = {str(k).strip().upper(): v
                     for k, v in zip(master_staff['Branch'], master_staff['RBM'])}
    branch_to_bdm = {str(k).strip().upper(): v
                     for k, v in zip(master_staff['Branch'], master_staff['BDM'])}

    # ── Load & clean feedback ───────────────────────────────────────────────
    df_fb = load_feedback_auto(fb_path)

    if 'Branch Name' in df_fb.columns:
        df_fb['Branch Name'] = df_fb['Branch Name'].astype(str).str.strip().str.upper()
        df_fb['Branch Name'] = df_fb['Branch Name'].replace('NAN', np.nan)

    def get_rbm(row):
        val = row.get('Staff Name')
        if pd.notna(val) and val in staff_to_rbm and pd.notna(staff_to_rbm[val]):
            return staff_to_rbm[val]
        b = row.get('Branch Name')
        if pd.notna(b) and b in branch_to_rbm and pd.notna(branch_to_rbm[b]):
            return branch_to_rbm[b]
        return row.get('RBM')

    def get_bdm(row):
        val = row.get('Staff Name')
        if pd.notna(val) and val in staff_to_bdm and pd.notna(staff_to_bdm[val]):
            return staff_to_bdm[val]
        b = row.get('Branch Name')
        if pd.notna(b) and b in branch_to_bdm and pd.notna(branch_to_bdm[b]):
            return branch_to_bdm[b]
        return row.get('BDM')

    df_fb['RBM'] = df_fb.apply(get_rbm, axis=1)
    df_fb['BDM'] = df_fb.apply(get_bdm, axis=1)
    df_fb['Rating'] = pd.to_numeric(df_fb['Rating'], errors='coerce').fillna(0.0)

    master_staff = master_staff[master_staff['Staff'].notna()]

    fb_staff_agg  = df_fb.groupby('Staff Name').agg(
        FB_Count=('Rating', 'count'), Rating_Avg=('Rating', 'mean')).reset_index()
    fb_branch_agg = df_fb.groupby('Branch Name').agg(
        FB_Count=('Rating', 'count'), Rating_Avg=('Rating', 'mean')).reset_index()
    fb_rbm_agg    = df_fb.groupby('RBM').agg(
        FB_Count=('Rating', 'count'), Rating_Avg=('Rating', 'mean')).reset_index()
    fb_bdm_agg    = df_fb.groupby('BDM').agg(
        FB_Count=('Rating', 'count'), Rating_Avg=('Rating', 'mean')).reset_index()

    # =========================================================================
    # STAFF WISE REPORT
    # =========================================================================
    staff_report = pd.DataFrame()
    staff_report['STAFF']  = master_staff['Staff'].values
    staff_report['BRANCH'] = master_staff['Branch'].values

    bill_cut_map   = dict(zip(staff_bill_cut['Staff Code'], staff_bill_cut['Staff Bills']))
    fb_count_map   = dict(zip(fb_staff_agg['Staff Name'], fb_staff_agg['FB_Count']))
    avg_rating_map = dict(zip(fb_staff_agg['Staff Name'], fb_staff_agg['Rating_Avg']))

    staff_report['TOTAL BILL CUT']   = master_staff['Staff Code'].map(bill_cut_map).fillna(0).astype(int).values
    staff_report['FEEDBACK COUNT']   = staff_report['STAFF'].map(fb_count_map).fillna(0).astype(int)
    staff_report['% CONVERSION']     = np.where(
        staff_report['TOTAL BILL CUT'] > 0,
        staff_report['FEEDBACK COUNT'] / staff_report['TOTAL BILL CUT'], 0.0)
    staff_report['RATINGS']          = staff_report['STAFF'].map(avg_rating_map).fillna(0.0).round(1)

    staff_report = staff_report.sort_values('FEEDBACK COUNT', ascending=False).reset_index(drop=True)
    staff_report = _add_total_row(
        staff_report, label_col='STAFF',
        sum_cols=['TOTAL BILL CUT', 'FEEDBACK COUNT'],
        avg_cols=['RATINGS'],
        pct_cols_raw=[('% CONVERSION', 'FEEDBACK COUNT', 'TOTAL BILL CUT')]
    )

    # =========================================================================
    # BRANCH WISE REPORT
    # =========================================================================
    branch_master = master_staff[['Branch']].drop_duplicates().dropna()
    branch_report = pd.DataFrame()
    branch_report['BRANCH'] = branch_master['Branch'].values

    b_bill_cut_map = dict(zip(branch_bill_cut['Branch'], branch_bill_cut['Branch Bills']))
    b_fb_count_map = {str(k).strip().upper(): v
                     for k, v in zip(fb_branch_agg['Branch Name'], fb_branch_agg['FB_Count'])}
    b_avg_map      = {str(k).strip().upper(): v
                     for k, v in zip(fb_branch_agg['Branch Name'], fb_branch_agg['Rating_Avg'])}

    branch_report['TOTAL BILL CUT']  = branch_report['BRANCH'].map(b_bill_cut_map).fillna(0).astype(int).values
    branch_report['FEEDBACK COUNT']  = branch_report['BRANCH'].str.strip().str.upper().map(b_fb_count_map).fillna(0).astype(int)
    branch_report['% CONVERSION']    = np.where(
        branch_report['TOTAL BILL CUT'] > 0,
        branch_report['FEEDBACK COUNT'] / branch_report['TOTAL BILL CUT'], 0.0)
    branch_report['RATINGS']         = branch_report['BRANCH'].str.strip().str.upper().map(b_avg_map).fillna(0.0).round(1)

    branch_report = branch_report.sort_values('% CONVERSION', ascending=False).reset_index(drop=True)
    branch_report = _add_total_row(
        branch_report, label_col='BRANCH',
        sum_cols=['TOTAL BILL CUT', 'FEEDBACK COUNT'],
        avg_cols=['RATINGS'],
        pct_cols_raw=[('% CONVERSION', 'FEEDBACK COUNT', 'TOTAL BILL CUT')]
    )

    # =========================================================================
    # RBM WISE REPORT
    # =========================================================================
    rbm_master = master_staff[['RBM']].drop_duplicates().dropna()
    rbm_report = pd.DataFrame()
    rbm_report['RBM'] = rbm_master['RBM'].values

    r_bill_cut_map = dict(zip(rbm_bill_cut['RBM'], rbm_bill_cut['RBM Bills']))
    r_fb_count_map = dict(zip(fb_rbm_agg['RBM'], fb_rbm_agg['FB_Count']))
    r_avg_map      = dict(zip(fb_rbm_agg['RBM'], fb_rbm_agg['Rating_Avg']))

    rbm_report['TOTAL BILL CUT']  = rbm_report['RBM'].map(r_bill_cut_map).fillna(0).astype(int).values
    rbm_report['FEEDBACK COUNT']  = rbm_report['RBM'].map(r_fb_count_map).fillna(0).astype(int)
    rbm_report['% CONVERSION']    = np.where(
        rbm_report['TOTAL BILL CUT'] > 0,
        rbm_report['FEEDBACK COUNT'] / rbm_report['TOTAL BILL CUT'], 0.0)
    rbm_report['RATINGS']         = rbm_report['RBM'].map(r_avg_map).fillna(0.0).round(1)

    rbm_report = rbm_report.sort_values('% CONVERSION', ascending=False).reset_index(drop=True)
    rbm_report = _add_total_row(
        rbm_report, label_col='RBM',
        sum_cols=['TOTAL BILL CUT', 'FEEDBACK COUNT'],
        avg_cols=['RATINGS'],
        pct_cols_raw=[('% CONVERSION', 'FEEDBACK COUNT', 'TOTAL BILL CUT')]
    )

    # =========================================================================
    # BDM WISE REPORT
    # =========================================================================
    bdm_master = master_staff[['BDM']].drop_duplicates().dropna()
    bdm_report = pd.DataFrame()
    bdm_report['BDM'] = bdm_master['BDM'].values

    d_bill_cut_map = dict(zip(bdm_bill_cut['BDM'], bdm_bill_cut['BDM Bills']))
    d_fb_count_map = dict(zip(fb_bdm_agg['BDM'], fb_bdm_agg['FB_Count']))
    d_avg_map      = dict(zip(fb_bdm_agg['BDM'], fb_bdm_agg['Rating_Avg']))

    bdm_report['TOTAL BILL CUT']  = bdm_report['BDM'].map(d_bill_cut_map).fillna(0).astype(int).values
    bdm_report['FEEDBACK COUNT']  = bdm_report['BDM'].map(d_fb_count_map).fillna(0).astype(int)
    bdm_report['% CONVERSION']    = np.where(
        bdm_report['TOTAL BILL CUT'] > 0,
        bdm_report['FEEDBACK COUNT'] / bdm_report['TOTAL BILL CUT'], 0.0)
    bdm_report['RATINGS']         = bdm_report['BDM'].map(d_avg_map).fillna(0.0).round(1)

    bdm_report = bdm_report.sort_values('% CONVERSION', ascending=False).reset_index(drop=True)
    bdm_report = _add_total_row(
        bdm_report, label_col='BDM',
        sum_cols=['TOTAL BILL CUT', 'FEEDBACK COUNT'],
        avg_cols=['RATINGS'],
        pct_cols_raw=[('% CONVERSION', 'FEEDBACK COUNT', 'TOTAL BILL CUT')]
    )

    # =========================================================================
    # Write to Excel with professional styling
    # =========================================================================
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        _write_styled_sheet(
            writer, rbm_report, sheet_name='RBM',
            title_label='RBM Wise Feedback Report',
            text_col_names=['RBM'],
            pct_col_names=['% CONVERSION'],
            rating_col_names=['RATINGS'],
            int_col_names=['TOTAL BILL CUT', 'FEEDBACK COUNT']
        )
        _write_styled_sheet(
            writer, bdm_report, sheet_name='BDM',
            title_label='BDM Wise Feedback Report',
            text_col_names=['BDM'],
            pct_col_names=['% CONVERSION'],
            rating_col_names=['RATINGS'],
            int_col_names=['TOTAL BILL CUT', 'FEEDBACK COUNT']
        )
        _write_styled_sheet(
            writer, branch_report, sheet_name='BRANCH',
            title_label='Branch Wise Feedback Report',
            text_col_names=['BRANCH'],
            pct_col_names=['% CONVERSION'],
            rating_col_names=['RATINGS'],
            int_col_names=['TOTAL BILL CUT', 'FEEDBACK COUNT']
        )
        _write_styled_sheet(
            writer, staff_report, sheet_name='STAFF',
            title_label='Staff Wise Feedback Report',
            text_col_names=['STAFF', 'BRANCH'],
            pct_col_names=['% CONVERSION'],
            rating_col_names=['RATINGS'],
            int_col_names=['TOTAL BILL CUT', 'FEEDBACK COUNT']
        )
        
        if 'Sheet' in writer.book.sheetnames and len(writer.book.sheetnames) > 1:
            del writer.book['Sheet']
        writer.book.active = 0


# ---------------------------------------------------------------------------
# SMS-Style Branch Summary Report builder (Section 02)
# ---------------------------------------------------------------------------
def process_monthly_report(sales_path: str, fb_path: str, output_path: str):
    """
    SMS Branch Feedback Report (Section 02).

    Sales Report  → unique Customer Mobile per Branch  = Total Bill Cut
    Feedback Report → Rating count per Branch Name      = Feedback Count
                      Rating mean  per Branch Name      = Avg Rating
    Output: branch-wise table sorted by % Conversion descending,
            styled to match the reference image (pink title, red header,
            green-shaded % Conversion column).
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── 1. Load sales data ───────────────────────────────────────────────
    df_sales_raw = load_sales_auto(sales_path)

    # Remove 'general' rows from key columns
    for col in ['RBM', 'BDM', 'Staff', 'Branch', 'Category', 'Designation']:
        if col in df_sales_raw.columns:
            df_sales_raw = df_sales_raw[~(df_sales_raw[col].astype(str).str.strip().str.lower() == 'general')]

    # Auto-detect date range
    date_range_str = ''
    for col in df_sales_raw.columns:
        if 'date' in str(col).lower():
            try:
                dates = pd.to_datetime(df_sales_raw[col], errors='coerce').dropna()
                if len(dates) > 0:
                    date_range_str = (
                        f"{dates.min().strftime('%d-%m-%Y')} - "
                        f"{dates.max().strftime('%d-%m-%Y')}")
                    break
            except Exception:
                pass

    BRANCH_COL = 'Branch'
    MOBILE_COL = 'Customer Mobile'

    avail = [c for c in [BRANCH_COL, MOBILE_COL] if c in df_sales_raw.columns]
    df_sales = df_sales_raw[avail].copy()
    del df_sales_raw
    gc.collect()

    # Unique mobile per branch → bill cut
    if BRANCH_COL in df_sales.columns and MOBILE_COL in df_sales.columns:
        df_unique = df_sales.drop_duplicates(subset=[MOBILE_COL])
        branch_bills = (df_unique.groupby(BRANCH_COL)
                        .size().reset_index(name='TOTAL BILL CUT'))
    else:
        branch_bills = pd.DataFrame(columns=[BRANCH_COL, 'TOTAL BILL CUT'])

    del df_sales
    gc.collect()

    # Normalise branch key for joining
    branch_bills['_key'] = (branch_bills[BRANCH_COL]
                            .astype(str).str.strip().str.upper())

    # ── 2. Load feedback data ────────────────────────────────────────────
    df_fb = load_feedback_auto(fb_path)

    FB_BRANCH = 'Branch Name'
    RATING_COL = 'Rating'

    if FB_BRANCH in df_fb.columns:
        df_fb[FB_BRANCH] = df_fb[FB_BRANCH].astype(str).str.strip().str.upper()
        df_fb[FB_BRANCH] = df_fb[FB_BRANCH].replace('NAN', np.nan)
        df_fb = df_fb.dropna(subset=[FB_BRANCH])

    if RATING_COL in df_fb.columns:
        df_fb[RATING_COL] = pd.to_numeric(df_fb[RATING_COL], errors='coerce')

    fb_agg = (df_fb.groupby(FB_BRANCH)
              .agg(FEEDBACK_COUNT=(RATING_COL, 'count'),
                   RATING=(RATING_COL, 'mean'))
              .reset_index())
    fb_agg.columns = ['_key', 'FEEDBACK COUNT', 'RATING']
    fb_agg['RATING'] = fb_agg['RATING'].round(1)

    # ── 3. Merge & build report ──────────────────────────────────────────
    merged = pd.merge(branch_bills, fb_agg, on='_key', how='outer')

    # Recover branch label: prefer sales name, fall back to feedback key
    merged['BRANCH'] = merged[BRANCH_COL].combine_first(merged['_key'])
    merged['TOTAL BILL CUT'] = merged['TOTAL BILL CUT'].fillna(0).astype(int)
    merged['FEEDBACK COUNT'] = merged['FEEDBACK COUNT'].fillna(0).astype(int)
    merged['RATING'] = merged['RATING'].fillna(0.0).round(1)
    merged['% CONVERSION'] = np.where(
        merged['TOTAL BILL CUT'] > 0,
        (merged['FEEDBACK COUNT'] / merged['TOTAL BILL CUT'] * 100).round(0).astype(int),
        0)

    report = (merged[['BRANCH', 'TOTAL BILL CUT', 'FEEDBACK COUNT',
                       'RATING', '% CONVERSION']]
              .sort_values('% CONVERSION', ascending=False)
              .reset_index(drop=True))

    # TOTAL row
    t_bill = report['TOTAL BILL CUT'].sum()
    t_fb   = report['FEEDBACK COUNT'].sum()
    valid_r = report.loc[report['RATING'] > 0, 'RATING']
    t_rat  = round(valid_r.mean(), 1) if len(valid_r) > 0 else 0.0
    t_conv = round(t_fb / t_bill * 100) if t_bill > 0 else 0

    report = pd.concat([report, pd.DataFrame([{
        'BRANCH': 'TOTAL',
        'TOTAL BILL CUT': t_bill,
        'FEEDBACK COUNT': t_fb,
        'RATING': t_rat,
        '% CONVERSION': t_conv
    }])], ignore_index=True)

    # ── 4. Write raw data then style with openpyxl ───────────────────────
    report.to_excel(output_path, index=False, engine='openpyxl')

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = 'SMS FEEDBACK REPORT'

    num_cols = len(report.columns)   # 5

    # Insert 2 top rows for title + generated-date
    ws.insert_rows(1)
    ws.insert_rows(1)

    title_text = 'SMS FEEDBACK REPORT'
    if date_range_str:
        title_text += f'  ({date_range_str})'

    # Row 1 — pink title (matches reference image)
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=num_cols)
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.font      = Font(name='Calibri', bold=True, size=13, color='000000')
    tc.fill      = PatternFill('solid', fgColor='FF99CC')
    tc.alignment = Alignment(horizontal='center', vertical='center')
    tc.border    = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000'))
    ws.row_dimensions[1].height = 26

    # Row 2 — generated date stamp
    dc = ws.cell(row=2, column=1,
                 value=f"Generated: {datetime.now().strftime('%d %b %Y  %I:%M %p')}")
    dc.font      = Font(name='Calibri', italic=True, size=9, color='7F7F7F')
    dc.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 14

    # Row 3 — red header (matches reference image)
    HEADER_ROW = 3
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx)
        cell.fill      = PatternFill('solid', fgColor='FF0000')
        cell.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000'))
    ws.row_dimensions[HEADER_ROW].height = 32

    # Column widths
    for i, w in enumerate([32, 18, 18, 12, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Conversion colour scale (green shades — reference image)
    def _conv_fill(pct):
        if pct >= 60:   return PatternFill('solid', fgColor='00B050')   # dark green
        if pct >= 40:   return PatternFill('solid', fgColor='92D050')   # medium green
        if pct >= 30:   return PatternFill('solid', fgColor='C6EFCE')   # light green
        if pct >= 20:   return PatternFill('solid', fgColor='FFEB9C')   # yellow-green
        return          PatternFill('solid', fgColor='FFC7CE')           # light red / low

    thin = Border(left=Side(style='thin', color='BFBFBF'),
                  right=Side(style='thin', color='BFBFBF'),
                  top=Side(style='thin', color='BFBFBF'),
                  bottom=Side(style='thin', color='BFBFBF'))
    thick = Border(left=Side(style='medium', color='1F3864'),
                   right=Side(style='medium', color='1F3864'),
                   top=Side(style='medium', color='1F3864'),
                   bottom=Side(style='medium', color='1F3864'))

    DATA_START = HEADER_ROW + 1
    total_rows = len(report)   # includes TOTAL row

    for r_offset, row_data in enumerate(report.itertuples(index=False), start=0):
        r_idx    = DATA_START + r_offset
        is_total = (r_offset == total_rows - 1)

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            if is_total:
                cell.fill   = PatternFill('solid', fgColor='D6E4F0')
                cell.font   = Font(name='Calibri', bold=True, size=11,
                                   color='1F3864')
                cell.border = thick
            else:
                cell.fill   = PatternFill('solid', fgColor='FFFFFF')
                cell.font   = Font(name='Calibri', size=10)
                cell.border = thin

            cell.alignment = Alignment(
                horizontal='left' if col_idx == 1 else 'center',
                vertical='center')

        # Colour-code % CONVERSION cell (col 5)
        conv_cell = ws.cell(row=r_idx, column=5)
        pct_val   = int(row_data[4])   # index 4 = '% CONVERSION'
        if is_total:
            conv_cell.value = f'{pct_val}%'
        else:
            conv_cell.fill  = _conv_fill(pct_val)
            conv_cell.value = f'{pct_val}%'
            conv_cell.font  = Font(name='Calibri', bold=True, size=10)

        ws.row_dimensions[r_idx].height = 18

    # Freeze header
    ws.freeze_panes = ws.cell(row=DATA_START, column=1)

    # Remove default Sheet if it crept in
    for sname in list(wb.sheetnames):
        if sname not in ('SMS FEEDBACK REPORT',) and len(wb.sheetnames) > 1:
            if sname == 'Sheet':
                del wb[sname]

    wb.save(output_path)
    gc.collect()


# ---------------------------------------------------------------------------
# Global error handlers — always return JSON so the frontend never receives HTML
# ---------------------------------------------------------------------------
@app.errorhandler(HTTPException)
def handle_http_exception(e):
    return jsonify({'status': 'error', 'message': f'{e.code} {e.name}: {e.description}'}), e.code


@app.errorhandler(Exception)
def handle_exception(e):
    traceback.print_exc()
    return jsonify({'status': 'error', 'message': str(e)}), 500


# ---------------------------------------------------------------------------
# Flask routes
# ---------------------------------------------------------------------------
@app.route('/health')
def health():
    """Health-check endpoint — Render pings this to confirm the app is up."""
    return jsonify({'status': 'ok'}), 200


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/process', methods=['POST'])
def process():
    """Generate the 4-sheet Feedback Analytics Report."""
    try:
        if 'sales_file' not in request.files or 'feedback_file' not in request.files:
            return jsonify({'status': 'error', 'message': 'Missing files'}), 400

        sales_file    = request.files['sales_file']
        feedback_file = request.files['feedback_file']

        if sales_file.filename == '' or feedback_file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400

        sales_path = os.path.join(UPLOAD_FOLDER, 'sales_r1.xlsx')
        fb_path    = os.path.join(UPLOAD_FOLDER, 'feedback_r1.xlsx')
        sales_file.save(sales_path)
        feedback_file.save(fb_path)

        timestamp   = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        out_name    = f'Feedback_Analytics_{timestamp}.xlsx'
        output_path = os.path.join(OUTPUT_FOLDER, out_name)

        process_reports(sales_path, fb_path, output_path)
        return jsonify({'status': 'success', 'filename': out_name})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/process-monthly', methods=['POST'])
def process_monthly():
    """Generate the standalone Monthly Branch Conversion Report."""
    try:
        if 'sales_file_m' not in request.files or 'feedback_file_m' not in request.files:
            return jsonify({'status': 'error', 'message': 'Missing files'}), 400

        sales_file    = request.files['sales_file_m']
        feedback_file = request.files['feedback_file_m']

        if sales_file.filename == '' or feedback_file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400

        sales_path = os.path.join(UPLOAD_FOLDER, 'sales_r2.xlsx')
        fb_path    = os.path.join(UPLOAD_FOLDER, 'feedback_r2.xlsx')
        sales_file.save(sales_path)
        feedback_file.save(fb_path)

        timestamp   = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        out_name    = f'SMS_Feedback_Report_{timestamp}.xlsx'
        output_path = os.path.join(OUTPUT_FOLDER, out_name)

        process_monthly_report(sales_path, fb_path, output_path)
        return jsonify({'status': 'success', 'filename': out_name})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/download/<filename>')
def download(filename):
    file_path = os.path.join(OUTPUT_FOLDER, filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    return "File not found", 404


if __name__ == '__main__':
    # Determine port: Use Render's PORT if deployed, otherwise strictly 7080 locally
    port = int(os.environ.get('PORT')) if os.environ.get('RENDER') else 7080
    print(f"Starting Feedback Report Portal server on port {port}...")
    # On Render, apps MUST bind to 0.0.0.0 to receive external traffic.
    # use_reloader=False to prevent Werkzeug spawning duplicate processes on production
    app.run(host='0.0.0.0', port=port, debug=False, use_reloader=False)
